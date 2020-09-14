from sqlalchemy import create_engine
from sqlalchemy import Column, Integer, String, Text, MetaData, Table,DateTime, Date , ForeignKey
from sqlalchemy.ext.declarative import declarative_base
from sqlalchemy.orm import sessionmaker, relationship
from datetime import datetime, timedelta
from PyQt5.QtWidgets import  QWidget,QTableWidgetItem
from PyQt5 import QtCore, QtGui, QtWidgets
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.table import Table, TableStyleInfo
import os
from datetime import datetime
import sys
sys.path.append('../Modelo/')
from farm import Clave,Salida,Farmaco,Historial
import pandas as pd
import pymysql
engine = create_engine('mysql+pymysql://root:wil99@localhost/prueba')
Session = sessionmaker(bind=engine)
session = Session()

class SubWindow(QWidget):
    def createSubWindow(self):
        self.setObjectName("InterFazConsulta")
        self.resize(1034, 687)
        self.setStyleSheet("*{\n"
"font-family:century gothic;\n"
"font-size: 12px;\n"
"}\n"
"#Frame2{\n"
"border-radius:30;\n"
"}\n"
"QLineEdit{\n"
"    border:none;\n"
"    border-bottom: 1px solid black;\n"
"}\n"
"QFrame{\n"
"border-radius: 60px;\n"
" background:#fefefe;\n"
"}\n"
"QLabel{\n"
"    background:transparent;\n"
"}\n"
"QCalendarWidget QAbstractItemView\n"
"{ \n"
"selection-background-color: #042944; \n"
"selection-color: white;\n"
"selection-border:10px solid red;\n"
"\n"
"}\n"
"QCalendarWidget QWidget \n"
"{\n"
"  color:grey;\n"
"}\n"
"QCalendarWidget QTableView{\n"
"border-width:0px;\n"
"background-color:lightgrey;\n"
"}\n"
"\n"
"QPushButton:hover{\n"
"background:#dea806;\n"
"}")
        self.frame_2 = QtWidgets.QFrame(self)
        self.frame_2.setGeometry(QtCore.QRect(20, 9, 991, 661))
        self.frame_2.setStyleSheet("\n"
" background:#fefefe;\n"
"\n"
"")
        self.frame_2.setFrameShape(QtWidgets.QFrame.StyledPanel)
        self.frame_2.setFrameShadow(QtWidgets.QFrame.Raised)
        self.frame_2.setObjectName("frame_2")
        self.TableConsulta = QtWidgets.QTableWidget(self.frame_2)
        self.TableConsulta.setGeometry(QtCore.QRect(200, 20, 771, 621))
        self.TableConsulta.setEditTriggers(QtWidgets.QAbstractItemView.NoEditTriggers)
        self.TableConsulta.setObjectName("TableConsulta")
        self.groupBox = QtWidgets.QGroupBox(self.frame_2)
        self.groupBox.setGeometry(QtCore.QRect(20, 20, 171, 621))
        self.groupBox.setStyleSheet("QGroupBox  {\n"
"    border: 1px solid gray;\n"
"    border-radius: 15px;\n"
"\n"
"}\n"
"\n"
"")
        self.groupBox.setTitle("")
        self.groupBox.setObjectName("groupBox")
        self.groupBox_4 = QtWidgets.QGroupBox(self.groupBox)
        self.groupBox_4.setGeometry(QtCore.QRect(10, 160, 151, 81))
        self.groupBox_4.setTitle("")
        self.groupBox_4.setObjectName("groupBox_4")
        self.comboboxTipoConsulta = QtWidgets.QComboBox(self.groupBox_4)
        self.comboboxTipoConsulta.setGeometry(QtCore.QRect(10, 20, 121, 22))
        self.comboboxTipoConsulta.setObjectName("comboboxTipoConsulta")
        self.comboboxTipoConsulta.addItem("")
        self.comboboxTipoConsulta.addItem("")
        self.groupBox_5 = QtWidgets.QGroupBox(self.groupBox)
        self.groupBox_5.setGeometry(QtCore.QRect(9, 30, 151, 101))
        self.groupBox_5.setStyleSheet("\n"
"\n"
"")
        self.groupBox_5.setTitle("")
        self.groupBox_5.setObjectName("groupBox_5")
        self.label_2 = QtWidgets.QLabel(self.groupBox_5)
        self.label_2.setGeometry(QtCore.QRect(30, 0, 47, 0))
        self.label_2.setObjectName("label_2")
        self.DateFechaPrimeroConsulta = QtWidgets.QDateEdit(self.groupBox_5)
        self.DateFechaPrimeroConsulta.setGeometry(QtCore.QRect(10, 10, 121, 22))
        self.DateFechaPrimeroConsulta.setAccelerated(False)
        self.DateFechaPrimeroConsulta.setProperty("showGroupSeparator", False)
        self.DateFechaPrimeroConsulta.setCalendarPopup(True)
        self.DateFechaPrimeroConsulta.setTimeSpec(QtCore.Qt.UTC)
        self.DateFechaPrimeroConsulta.setDate(QtCore.QDate(2020, 1, 1))
        self.DateFechaPrimeroConsulta.setObjectName("DateFechaPrimeroConsulta")
        self.DateFechaSegundoConsulta = QtWidgets.QDateEdit(self.groupBox_5)
        self.DateFechaSegundoConsulta.setGeometry(QtCore.QRect(10, 60, 121, 22))
        self.DateFechaSegundoConsulta.setAccelerated(False)
        self.DateFechaSegundoConsulta.setProperty("showGroupSeparator", False)
        self.DateFechaSegundoConsulta.setCalendarPopup(True)
        self.DateFechaSegundoConsulta.setTimeSpec(QtCore.Qt.UTC)
        self.DateFechaSegundoConsulta.setDate(QtCore.QDate(2020, 1, 1))
        self.DateFechaSegundoConsulta.setObjectName("DateFechaSegundoConsulta")
        self.label_3 = QtWidgets.QLabel(self.groupBox_5)
        self.label_3.setGeometry(QtCore.QRect(40, 30, 41, 16))
        self.label_3.setStyleSheet("background:transparent;")
        self.label_3.setObjectName("label_3")
        self.btnExportarConsulta = QtWidgets.QPushButton(self.groupBox)
        self.btnExportarConsulta.setGeometry(QtCore.QRect(10, 560, 141, 31))
        self.btnExportarConsulta.setStyleSheet("QPushButton{\n"
"border-radius:15px;\n"
"background:#ffc001;\n"
"\n"
"}\n"
"QPushButton:hover{\n"
"background:#dea806;\n"
"}\n"
"")
        self.btnExportarConsulta.setObjectName("btnExportarConsulta")
        self.groupBox_6 = QtWidgets.QGroupBox(self.groupBox)
        self.groupBox_6.setGeometry(QtCore.QRect(10, 380, 151, 71))
        self.groupBox_6.setStyleSheet("")
        self.groupBox_6.setTitle("")
        self.groupBox_6.setObjectName("groupBox_6")
        self.label_6 = QtWidgets.QLabel(self.groupBox_6)
        self.label_6.setGeometry(QtCore.QRect(30, 0, 47, 0))
        self.label_6.setObjectName("label_6")
        self.comboboxSETConsulta = QtWidgets.QComboBox(self.groupBox_6)
        self.comboboxSETConsulta.setGeometry(QtCore.QRect(10, 10, 121, 22))
        self.comboboxSETConsulta.setObjectName("comboboxSETConsulta")
        self.comboboxSETConsulta.addItem("")
        self.comboboxSETConsulta.addItem("")
        self.comboboxSETConsulta.addItem("")
        self.radioButton = QtWidgets.QRadioButton(self.groupBox)
        self.radioButton.setGeometry(QtCore.QRect(34, 10, 71, 20))
        self.radioButton.setObjectName("radioButton")
        self.radioButton_2 = QtWidgets.QRadioButton(self.groupBox)
        self.radioButton_2.setGeometry(QtCore.QRect(33, 140, 70, 17))
        self.radioButton_2.setObjectName("radioButton_2")
        self.groupBox_2 = QtWidgets.QGroupBox(self.groupBox)
        self.groupBox_2.setGeometry(QtCore.QRect(10, 280, 151, 61))
        self.groupBox_2.setTitle("")
        self.groupBox_2.setObjectName("groupBox_2")
        self.comboboxExistenciaConsulta = QtWidgets.QComboBox(self.groupBox_2)
        self.comboboxExistenciaConsulta.setGeometry(QtCore.QRect(10, 10, 121, 22))
        self.comboboxExistenciaConsulta.setObjectName("comboboxExistenciaConsulta")
        self.comboboxExistenciaConsulta.addItem("")
        self.comboboxExistenciaConsulta.addItem("")
        self.comboboxExistenciaConsulta.addItem("")
        self.radioButton_3 = QtWidgets.QRadioButton(self.groupBox)
        self.radioButton_3.setGeometry(QtCore.QRect(30, 250, 81, 17))
        self.radioButton_3.setObjectName("radioButton_3")
        self.label_4 = QtWidgets.QLabel(self.groupBox)
        self.label_4.setGeometry(QtCore.QRect(29, 350, 101, 16))
        self.label_4.setObjectName("label_4")
        self.lineFilasConsulta = QtWidgets.QLineEdit(self.groupBox)
        self.lineFilasConsulta.setEnabled(False)
        self.lineFilasConsulta.setGeometry(QtCore.QRect(30, 480, 81, 21))
        self.lineFilasConsulta.setObjectName("lineFilasConsulta")
        self.label_5 = QtWidgets.QLabel(self.groupBox)
        self.label_5.setGeometry(QtCore.QRect(25, 460, 91, 16))
        self.label_5.setObjectName("label_5")
        self.btnConsultarConsulta = QtWidgets.QPushButton(self.groupBox)
        self.btnConsultarConsulta.setGeometry(QtCore.QRect(10, 510, 141, 31))
        self.btnConsultarConsulta.setStyleSheet("QPushButton{\n"
"border-radius:15px;\n"
"background:#ffc001;\n"
"\n"
"}\n"
"QPushButton:hover{\n"
"background:#dea806;\n"
"}\n"
"")
        self.btnConsultarConsulta.setObjectName("btnConsultarConsulta")
        self.label = QtWidgets.QLabel(self.frame_2)
        self.label.setGeometry(QtCore.QRect(40, 10, 61, 20))
        self.label.setObjectName("label")

        self.retranslateUi()
        self.radioButton_3.clicked.connect(self.groupBox_6.hide)
        self.radioButton.clicked.connect(self.groupBox_6.show)
        self.radioButton_2.clicked.connect(self.groupBox_6.show)
        self.radioButton.clicked.connect(self.label_4.show)
        self.radioButton_2.clicked.connect(self.label_4.show)
        self.radioButton_3.clicked.connect(self.label_4.hide)
        QtCore.QMetaObject.connectSlotsByName(self)

        listaMeses31 = {1,3,5,7,8,10,12}
        ListaMeses30 = {4,6,9,11}

        #marcando el radiobutton de fechas
        self.radioButton.setChecked(True)
        self.tipo = ''
        #poniendo la fecha actual y un mes despues
        Fecha1 = datetime.now()
        self.DateFechaSegundoConsulta.setDate(Fecha1)
        #dependiendo si el mes tiene 31 dias o 30 o 28 es la cantidad que se suma
        if Fecha1.month in listaMeses31:
                dias = 31
        elif Fecha1.month in ListaMeses30:
                dias = 30
        elif Fecha1.month == 2:
                dias = 28
        Fecha2 = Fecha1 - timedelta(days=dias)
        self.DateFechaPrimeroConsulta.setDate(Fecha2)

        

        

        self.consultar()
        self.btnConsultarConsulta.clicked.connect(self.consultar)
        self.TipoConsultaTipoConsultaTipoConsulta = self.comboboxSETConsulta.currentIndex()
        
        self.btnExportarConsulta.clicked.connect(self.exportQueryEx)
       
        #
        #
        #SI QUIEREN CONSULTA POR NUMERO DE PEDIDO
        #AGREGAR EL CAMBPO Npedido
        #SOLO EN LAS SALIDAS (CUANDO self.TipoConsulta == 0 )
        #
        #
        #
    def consultar(self):
        if self.radioButton.isChecked():
                self.consultaXfechas()
        if self.radioButton_2.isChecked():
                self.consultaxTipo()
        if self.radioButton_3.isChecked():
                self.consultaxExistencia()



        #esta es la funcion de la consulta por fechas
    def consultaXfechas(self):
        Date = self.DateFechaPrimeroConsulta.date()
        vFecha = str(Date.toPyDate())
        Date = self.DateFechaSegundoConsulta.date()
        VFecha2 = str(Date.toPyDate())
        self.TipoConsulta = self.comboboxSETConsulta.currentIndex()
        # salidas =0  entradas =1 y totales = 2
        if self.TipoConsulta == 0:
                #opcional = Salida
                #La consulta de toda la tabla
                Query = session.query(Salida.idSalida,Salida.clave_corta,Clave.descripcion,Clave.presentacion,Salida.cantidadSal,Salida.Caducidad,Salida.FechaPedido,Salida.fechaEntrega,Salida.area,Salida.lote,Salida.numero_pedido).join(Clave).filter(Salida.FechaPedido.between(vFecha,VFecha2)).all()
                #Query = pd.read_sql('SELECT * FROM Salida WHERE (FechaPedido BETWEEN '+vFecha+' AND '+VFecha2+')',engine)
                self.TableConsulta.setColumnCount(11)
                #el encabezado de la tabla
                self.TableConsulta.setHorizontalHeaderLabels(['id Salida','Clave','Descripcion','Presentacion','Cantidad','Caducidad','Fecha pedido','Fecha entrega','Destino','Lote','N pedido'])
                #aqui le indicamos cuantas filas tendra nuestra tabla
                self.filas = len(Query)
                self.TableConsulta.setRowCount(self.filas)
                self.fillTableQuery(Query)
                
                
        if self.TipoConsulta == 1:
                Query = session.query(Historial.idFarmaco,Historial.clave_corta,Clave.descripcion,Clave.presentacion,Historial.cantidad,Historial.caducidad,Historial.area,Historial.origen,Historial.fechaIngreso,Historial.lote).join(Clave).filter(Historial.fechaIngreso.between(vFecha,VFecha2)).all()
                self.TableConsulta.setColumnCount(10)
                self.TableConsulta.setHorizontalHeaderLabels(['id Farmaco','Clave','Descripcion','Presentacion','Cantidad','Caducidad','Area almacen','Origen','Fecha ingreso','Lote'])
                self.filas = len(Query)
                self.TableConsulta.setRowCount(self.filas)
                self.fillTableQuery(Query)
        
        if self.TipoConsulta == 2:
                Query = session.query(Salida.idSalida,Salida.clave_corta,Clave.descripcion,Clave.presentacion,Salida.cantidadSal,Salida.Caducidad,Salida.area,Salida.fechaEntrega,Salida.lote).join(Clave).filter(Salida.FechaPedido.between(vFecha,VFecha2)).all()
                Query1 = session.query(Historial.idFarmaco,Historial.clave_corta,Clave.descripcion,Clave.presentacion,Historial.cantidad,Historial.caducidad,Historial.origen,Historial.fechaIngreso,Historial.lote).join(Clave).filter(Historial.fechaIngreso.between(vFecha,VFecha2)).all()
                self.TableConsulta.setColumnCount(10)
                self.TableConsulta.setHorizontalHeaderLabels(['id','Clave','Descripcion','Presentacion','Cantidad','Caducidad','Origen/Destino','Fecha Salida/Entrega','Lote','Entrada o Salida'])
                self.filas = len(Query) + len(Query1)
                self.TableConsulta.setRowCount(self.filas)
                self.maxCol = 9
                self.rows = 0
                self.SoE = 'Salida'              
                self.fillTableQuery(Query)
                Query = session.query(Historial.idFarmaco,Historial.clave_corta,Clave.descripcion,Clave.presentacion,Historial.cantidad,Historial.caducidad,Historial.origen,Historial.fechaIngreso,Historial.lote).join(Clave).filter(Historial.fechaIngreso.between(vFecha,VFecha2)).all()
                self.SoE = 'Entrada'
                self.fillTableQuery(Query)               

                
        #esta es la consulta por tipo
    def consultaxTipo(self):
        #este es el tipo de si medicina o material de curacion
        self.Por_Tipo = self.comboboxTipoConsulta.currentIndex()
        #es el tipo de salida , entra o por ambas
        self.TipoConsulta = self.comboboxSETConsulta.currentIndex()
        if self.TipoConsulta == 0:
                Query = session.query(Salida.idSalida,Salida.clave_corta,Clave.descripcion,Clave.presentacion,Salida.cantidadSal,Salida.Caducidad,Salida.FechaPedido,Salida.fechaEntrega,Salida.area,Salida.lote,Salida.numero_pedido).join(Clave).filter(Clave.tipo == self.Por_Tipo).all()
                self.TableConsulta.setColumnCount(11)
                self.TableConsulta.setHorizontalHeaderLabels(['Id Salida','Clave','Descripcion','Presentacion','Cantidad','Caducidad','Fecha Pedido','Fecha Salida','Destino','Lote','N pedido'])
                self.filas = len(Query)
                self.TableConsulta.setRowCount(self.filas)
                self.fillTableQuery(Query)
        if self.TipoConsulta == 1:
                Query = session.query(Historial.idFarmaco,Historial.clave_corta,Clave.descripcion,Clave.presentacion,Historial.cantidad,Historial.caducidad,Historial.fechaIngreso,Historial.area,Historial.origen,Historial.lote).join(Clave).filter(Clave.tipo == self.Por_Tipo).all()
                self.TableConsulta.setColumnCount(10)
                self.TableConsulta.setHorizontalHeaderLabels(['Id ','Clave','Descripcion','Presentacion','Cantidad','Caducidad','Fecha Ingreso','Area Almacen','Origen','Lote'])
                self.filas = len(Query)
                self.TableConsulta.setRowCount(self.filas)
                self.fillTableQuery(Query)

        if self.TipoConsulta == 2:
                #hay algo rato en tipo
                Query = session.query(Salida.idSalida,Salida.clave_corta,Clave.descripcion,Clave.presentacion,Salida.cantidadSal,Salida.Caducidad,Salida.fechaEntrega,Salida.area,Salida.lote).join(Clave).filter(Clave.tipo == self.Por_Tipo).all()
                Query1 = session.query(Historial.idFarmaco,Historial.clave_corta,Clave.descripcion,Clave.presentacion,Historial.cantidad,Historial.caducidad,Historial.fechaIngreso,Historial.origen,Historial.lote).join(Clave).filter(Clave.tipo == self.Por_Tipo).all()
                self.TableConsulta.setColumnCount(10)
                self.TableConsulta.setHorizontalHeaderLabels(['Id','Clave','Descripcion','Presentacion','Cantidad','Caducidad','Fecha Salida/Entrada','Destino/Origen','Lote','Tipo'])
                self.filas = len(Query) + len(Query1)
                self.TableConsulta.setRowCount(self.filas)
                self.maxCol = 9
                self.rows = 0
                if self.Por_Tipo == 0:
                        self.SoE = 'Medicina'    
                else:
                        self.SoE = 'Ma-Curacion'          
                self.fillTableQuery(Query)
                Query = session.query(Historial.idFarmaco,Historial.clave_corta,Clave.descripcion,Clave.presentacion,Historial.cantidad,Historial.caducidad,Historial.fechaIngreso,Historial.origen,Historial.lote).join(Clave).filter(Clave.tipo == self.Por_Tipo).all()
                
                self.fillTableQuery(Query)

    def consultaxExistencia(self):
        self.TipoConsulta = self.comboboxExistenciaConsulta.currentIndex()
        if self.TipoConsulta == 0 :
                Query = session.query(Farmaco.idFarmaco,Farmaco.clave_corta,Clave.descripcion,Clave.presentacion,Farmaco.cantidad,Farmaco.caducidad,Farmaco.area,Farmaco.origen,Farmaco.lote,Clave.tipo).join(Clave).filter(Clave.tipo == self.TipoConsulta).all()
                self.TableConsulta.setColumnCount(10)
                self.tipo = 'Medicina'
                #pone false en tipo
                self.TableConsulta.setHorizontalHeaderLabels(['id Farmaco','Clave','Descripcion','Presentacion','Cantidad','Caducidad','Area almacen','Origen','Lote','Tipo'])
                self.filas = len(Query)
                self.TableConsulta.setRowCount(self.filas)
                self.fillTableQuery(Query)
                for a in range(self.filas):
                        self.TableConsulta.setItem(a,9,QTableWidgetItem(self.tipo))

        if self.TipoConsulta == 1:
                Query = session.query(Farmaco.idFarmaco,Farmaco.clave_corta,Clave.descripcion,Clave.presentacion,Farmaco.cantidad,Farmaco.caducidad,Farmaco.area,Farmaco.origen,Farmaco.lote,Clave.tipo).join(Clave).filter(Clave.tipo == self.TipoConsulta).all()
                self.TableConsulta.setColumnCount(10)
                self.tipo = 'Ma-Curacion'
                #pone false en tipo
                self.TableConsulta.setHorizontalHeaderLabels(['id Farmaco','Clave','Descripcion','Presentacion','Cantidad','Caducidad','Area almacen','Origen','Lote','Tipo'])
                self.filas = len(Query)
                self.TableConsulta.setRowCount(self.filas)
                self.fillTableQuery(Query)
                for a in range(self.filas):
                        self.TableConsulta.setItem(a,9,QTableWidgetItem(self.tipo))

        if self.TipoConsulta == 2:
                Query = session.query(Farmaco.idFarmaco,Farmaco.clave_corta,Clave.descripcion,Clave.presentacion,Farmaco.cantidad,Farmaco.caducidad,Farmaco.area,Farmaco.origen,Farmaco.lote).join(Clave).filter(Clave.tipo == 0).all()
                Query1 = session.query(Farmaco.idFarmaco,Farmaco.clave_corta,Clave.descripcion,Clave.presentacion,Farmaco.cantidad,Farmaco.caducidad,Farmaco.area,Farmaco.origen,Farmaco.lote).join(Clave).filter(Clave.tipo == 1).all()
                self.TableConsulta.setColumnCount(10)
                self.tipo = 'Ma-Curacion'
                #pone false en tipo
                self.TableConsulta.setHorizontalHeaderLabels(['id Farmaco','Clave','Descripcion','Presentacion','Cantidad','Caducidad','Area almacen','Origen','Lote','Tipo'])
                self.filas = len(Query) + len(Query1)
                self.TableConsulta.setRowCount(self.filas)
                self.maxCol = 9
                self.rows = 0
                self.SoE = 'Medicina'              
                self.fillTableQuery(Query)
                Query = session.query(Farmaco.idFarmaco,Farmaco.clave_corta,Clave.descripcion,Clave.presentacion,Farmaco.cantidad,Farmaco.caducidad,Farmaco.area,Farmaco.origen,Farmaco.lote).join(Clave).filter(Clave.tipo == 1).all()
                self.SoE = 'Ma-Curacion'
                self.fillTableQuery(Query)

            
    def fillTableQuery(self,Query):
        self.data = Query
        if self.TipoConsulta == 0 or self.TipoConsulta == 1:
            rows = 0
            col = 0
            #hacemos un for anidado en el cual el primero se encarga de las filas de la  consulta
            #y el segundo de las columnas de cada fila
            #a contiene una fila y b contiene una columna de la fila
            for a in Query:
                for b in a:
                    self.TableConsulta.setItem(rows,col,QTableWidgetItem(str(b)))
                    # if self.tipo !='':
                    #         self.TableConsulta.setItem(rows,col,QTableWidgetItem(str(self.tipo)))
                    col = col + 1
                col = 0
                rows = rows + 1
            self.lineFilasConsulta.setText(str(self.filas))
            
        if self.TipoConsulta == 2:
            #aqui no declaro las rows (filas) en 0 por que son dos consultas que pasan por aqui. si se llegara a poner en 0 ensimaria los datos y no se mostarian todos
            #declaro las filas arriba en el tipoconsulta = 2
            col = 0
            for a in Query:
                for b in a:
                    self.TableConsulta.setItem(self.rows,col,QTableWidgetItem(str(b)))
                    col = col + 1
                    #QUITAR EL ESTATICO PARA MANDAR MAS COLUMNAS DESDE LAS FUNCIONES
                    if col == self.maxCol:
                            #ponemos si la consulta proviene de Salidas o de Entradas(Historial) y colocamos el resultado en la celda 0,7
                            self.TableConsulta.setItem(self.rows,self.maxCol,QTableWidgetItem(self.SoE))
                col = 0
                self.rows = self.rows + 1
            self.lineFilasConsulta.setText(str(self.filas))

    def exportQueryEx(self):
        now = datetime.now()
        outfilename = 'ExportConsul-{0}-{1}-{2}-{3}-{4}.xlsx'.format(now.year, now.month, now.day, now.hour, now.second)
        outfilepath = os.path.join(os.path.expanduser("~"), "Documents/Reportes", outfilename)

        rows = self.TableConsulta.rowCount()
        cols = self.TableConsulta.columnCount()
        df = pd.DataFrame()
        headers = []
        for i in range(rows):
            for j in range(cols):
                df.loc[i,j] = self.TableConsulta.item(i,j).text()
        for x in range(self.TableConsulta.columnCount()):
            headers.append(self.TableConsulta.horizontalHeaderItem(x).text())
        df.columns = headers
        wb = Workbook()
        ws = wb.active
        name_cols = ["A", "B", "C", "D", "E", "F", "G", "H", "I", "J", "K", "L"]

        for row in dataframe_to_rows(df, index=False, header=True):
            ws.append(row)

        rows = ws.max_row
        a = name_cols[ws.max_column - 1]
        for x in range(ws.max_column):
            ws.column_dimensions[name_cols[x]].width = 20

        tab = Table(displayName="Table1", ref="A1:{0}{1}".format(a, rows))
        style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False,
                               showLastColumn=False, showRowStripes=True, showColumnStripes=True)
        tab.tableStyleInfo = style
        ws.add_table(tab)

        wb.save(outfilepath)
        os.startfile(outfilepath)



    def retranslateUi(self):
        _translate = QtCore.QCoreApplication.translate
        self.setWindowTitle(_translate("InterFazConsulta", "Consultas"))
        self.comboboxTipoConsulta.setItemText(0, _translate("InterFazConsulta", "Medicina"))
        self.comboboxTipoConsulta.setItemText(1, _translate("InterFazConsulta", "Material/Curacion"))
        self.label_2.setText(_translate("InterFazConsulta", "TextLabel"))
        self.DateFechaPrimeroConsulta.setDisplayFormat(_translate("InterFazConsulta", "dd/MM/yyyy"))
        self.DateFechaSegundoConsulta.setDisplayFormat(_translate("InterFazConsulta", "dd/MM/yyyy"))
        self.label_3.setText(_translate("InterFazConsulta", "______"))
        self.btnExportarConsulta.setText(_translate("InterFazConsulta", "Exportar"))
        self.label_6.setText(_translate("InterFazConsulta", "TextLabel"))
        self.comboboxSETConsulta.setItemText(0, _translate("InterFazConsulta", "Salidas"))
        self.comboboxSETConsulta.setItemText(1, _translate("InterFazConsulta", "Entradas"))
        self.comboboxSETConsulta.setItemText(2, _translate("InterFazConsulta", "Ambas"))
        self.radioButton.setText(_translate("InterFazConsulta", "Fechas"))
        self.radioButton_2.setText(_translate("InterFazConsulta", "Tipo"))
        self.comboboxExistenciaConsulta.setItemText(0, _translate("InterFazConsulta", "Medicina"))
        self.comboboxExistenciaConsulta.setItemText(1, _translate("InterFazConsulta", "Material/Curacion"))
        self.comboboxExistenciaConsulta.setItemText(2, _translate("InterFazConsulta", "Ambas"))
        self.radioButton_3.setText(_translate("InterFazConsulta", "Existencia"))
        self.label_4.setText(_translate("InterFazConsulta", "Tipo de consulta"))
        self.label_5.setText(_translate("InterFazConsulta", "Numero de filas"))
        self.btnConsultarConsulta.setText(_translate("InterFazConsulta", "Consultar"))
        self.label.setText(_translate("InterFazConsulta", "Consultas"))


if __name__ == "__main__":
    import sys
    app = QtWidgets.QApplication(sys.argv)
    InterFazConsulta = QtWidgets.QWidget()
    ui = SubWindow()
    ui.createSubWindow()
    InterFazConsulta.show()
    sys.exit(app.exec_())
